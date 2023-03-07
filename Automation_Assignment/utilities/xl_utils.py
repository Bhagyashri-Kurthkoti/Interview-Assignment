import os.path
import openpyxl

filename=os.path.abspath(os.path.join(os.getcwd(), "..", "TestData", "excel_sample_data.xlsx"))
def getRowCount(filename, sheetName):
    workbook = openpyxl.load_workbook(filename)
    sheetname = workbook[sheetName]
    return sheetname.max_row


def getColumnCount(filename, sheetName):
    workbook = openpyxl.load_workbook(filename)
    sheetname = workbook[sheetName]
    return sheetname.max_column


def readData(filename, sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(filename)
    sheetname = workbook[sheetName]
    return sheetname.cell(row=rownum,column=colnum).value


def writeData(filename, sheetName, rownum, colnum, data):
    workbook = openpyxl.load_workbook(filename)
    sheetname = workbook[sheetName]
    sheetname.cell(row=rownum, col=colnum).value = data
    workbook.save(filename)
